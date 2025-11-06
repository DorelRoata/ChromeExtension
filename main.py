import os
import sys
import argparse

import time
import queue
import logging
import threading
import webbrowser
import subprocess
import tkinter as tk
from tkinter import simpledialog, messagebox, scrolledtext
import tkinter.font as tkFont
from datetime import datetime, date
from openpyxl import load_workbook
from flask import Flask, request, jsonify
from flask_cors import CORS

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Memory Management
def set_window_icon_safe(window, icon_ico_path=None, icon_png_path=None):
    """Set window icon with proper cleanup to prevent memory leaks"""
    try:
        # Try ICO first (Windows native format)
        if icon_ico_path and os.path.exists(icon_ico_path):
            window.iconbitmap(icon_ico_path)
            return

        # Fallback to PNG - store reference on window to prevent garbage collection
        if icon_png_path and os.path.exists(icon_png_path):
            icon = tk.PhotoImage(file=icon_png_path)
            window.iconphoto(True, icon)
            # Store reference on window itself, not global
            window._icon_ref = icon
    except Exception as e:
        logger.warning(f"Could not set window icon: {e}")

# Error Handling
def handle_error(exception, context="", show_user=True, parent_window=None, log_level="error"):
    """Centralized error handling with user-friendly messages"""
    # Generate user-friendly message based on exception type
    if isinstance(exception, FileNotFoundError):
        user_msg = f"File not found: {context}\n\nCheck if the file exists and you have permission to access it."
    elif isinstance(exception, PermissionError):
        user_msg = f"Access denied: {context}\n\nPlease close the file in Excel and try again."
    elif isinstance(exception, TimeoutError):
        user_msg = f"Operation timed out: {context}\n\nTry again or check your internet connection."
    elif "Excel" in str(exception) or "openpyxl" in str(exception):
        user_msg = f"Excel error: {context}\n\nMake sure the Excel file is not open and try again."
    elif isinstance(exception, ValueError):
        user_msg = f"Invalid data: {context}\n\nCheck your input and try again."
    else:
        user_msg = f"Error: {context}\n\n{str(exception)}"

    # Log technical error
    if log_level == "error":
        logger.error(f"{context}: {exception}", exc_info=True)
    elif log_level == "warning":
        logger.warning(f"{context}: {exception}")
    else:
        logger.info(f"{context}: {exception}")

    # Show user-friendly message
    if show_user:
        try:
            if parent_window is not None:
                messagebox.showerror("Error", user_msg, parent=parent_window)
            else:
                messagebox.showerror("Error", user_msg)
        except Exception:
            print(user_msg, file=sys.stderr)

# Global configuration
SERVER_PORT = 5000
DATA_QUEUE = queue.Queue(maxsize=50)  # Limit queue size to prevent unbounded growth
FLASK_APP = None
TABS_TO_CLOSE = set()  # Track tabs that should be closed
REGISTERED_TABS = {}  # Track all open tabs {tabId: {'url': url, 'timestamp': time.time()}}

# Schema fields
FIELDS = [
    'ACI #', 'MFR Part #', 'MFR', 'Description', 'QTY', 'Per',
    'Vendor', 'Vendor Part #', 'Legacy', 'Unit Price', 'Date',
    'Change %', 'Last Updated Price', 'Last Updated Date', 'Price History'

]

VENDOR_URLS = {
    'grainger': 'https://www.grainger.com/product/{}/',
    'mcmaster-carr': 'https://www.mcmaster.com/{}/',
    'mcmaster': 'https://www.mcmaster.com/{}/',
    'festo': 'https://www.festo.com/us/en/a/{}',
    'zoro': 'https://www.zoro.com/i/{}/'
}

# Excel column number formats (0-indexed)
COLUMN_FORMATS = {
    9: r'_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)',  # Unit Price (col 10)
    10: 'mm-dd-yy',  # Date (col 11)
    12: r'_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)',  # Last Updated Price (col 13)
    13: 'mm-dd-yy',  # Last Updated Date (col 14)
}

#
# FLASK SERVER
#
def cleanup_stale_tabs(max_age_seconds=1800):
    """Remove tabs that haven't been cleaned up properly (older than 30 minutes)"""
    current_time = time.time()
    stale_tabs = [tab_id for tab_id, info in REGISTERED_TABS.items()
                  if current_time - info['timestamp'] > max_age_seconds]
    for tab_id in stale_tabs:
        REGISTERED_TABS.pop(tab_id, None)
        TABS_TO_CLOSE.discard(tab_id)
        logger.info(f"Cleaned up stale tab {tab_id}")

def clear_stale_data():
    """Clear stale data before new scraping operation"""
    # Clear data queue
    while not DATA_QUEUE.empty():
        try:
            DATA_QUEUE.get_nowait()
        except queue.Empty:
            break

    # Clear tabs to close
    TABS_TO_CLOSE.clear()

    # Remove old registered tabs (older than 5 seconds)
    current_time = time.time()
    stale_tabs = [tab_id for tab_id, info in REGISTERED_TABS.items()
                  if current_time - info['timestamp'] > 5]
    for tab_id in stale_tabs:
        REGISTERED_TABS.pop(tab_id, None)

    logger.info("Cleared stale data and old tab registrations")

def create_flask_app():
    app = Flask(__name__)
    CORS(app)
    
    @app.route('/ping', methods=['GET'])
    def ping():
        return jsonify({"status": "ok", "timestamp": time.time()})

    @app.route('/scrape', methods=['POST'])
    def receive_scrape():
        try:
            data = request.json
            logger.info(f"Received data from extension: {data.get('vendor')} - {data.get('partNumber')} (Tab ID: {data.get('tabId')})")
            try:
                DATA_QUEUE.put(data, block=False)
            except queue.Full:
                logger.warning("Queue full, discarding oldest item")
                try:
                    DATA_QUEUE.get_nowait()  # Remove oldest
                except queue.Empty:
                    pass
                DATA_QUEUE.put(data, block=False)  # Add new
            return jsonify({"status": "success"}), 200
        except Exception as e:
            logger.error(f"Error receiving data: {e}")
            return jsonify({"status": "error", "message": str(e)}), 500

    @app.route('/register-tab', methods=['POST'])
    def register_tab():
        """Register a tab when it opens"""
        try:
            data = request.json
            tab_id = data.get('tabId')
            url = data.get('url', 'unknown')
            if tab_id:
                REGISTERED_TABS[tab_id] = {'url': url, 'timestamp': time.time()}
                logger.info(f"Tab {tab_id} registered: {url}")

                # Clean up stale tabs older than 30 minutes
                cleanup_stale_tabs()
            return jsonify({"status": "success"}), 200
        except Exception as e:
            logger.error(f"Error registering tab: {e}")
            return jsonify({"status": "error"}), 500

    @app.route('/should-close/<int:tab_id>', methods=['GET'])
    def should_close_tab(tab_id):
        """Check if a tab should be closed"""
        should_close = tab_id in TABS_TO_CLOSE
        if should_close:
            TABS_TO_CLOSE.discard(tab_id)  # Remove after checking
            REGISTERED_TABS.pop(tab_id, None)  # Clean up registration
            logger.info(f"Signaling tab {tab_id} to close")
        return jsonify({"shouldClose": should_close})

    @app.route('/tab-closed', methods=['POST'])
    def tab_closed():
        """Clean up when a tab is manually closed or times out"""
        try:
            data = request.get_json(silent=True)
            if data:
                tab_id = data.get('tabId')
                if tab_id:
                    TABS_TO_CLOSE.discard(tab_id)  # Remove from close queue
                    REGISTERED_TABS.pop(tab_id, None)  # Remove from registered tabs
                    logger.info(f"Tab {tab_id} closed/cleaned up")
            return '', 204  # No content response
        except Exception as e:
            logger.error(f"Error handling tab closure: {e}")
            return '', 204  # Still return success to avoid client errors

    return app

def start_flask_server():
    """Start Flask server in background thread"""
    global FLASK_APP

    # Check if port is already in use
    import socket
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    result = sock.connect_ex(('localhost', SERVER_PORT))
    sock.close()

    if result == 0:
        logger.warning(f"Port {SERVER_PORT} is already in use. Another instance may be running.")
        from tkinter import messagebox
        response = messagebox.askyesno(
            "Port Already in Use",
            f"Port {SERVER_PORT} is already in use.\n\n"
            "This usually means another instance is running.\n\n"
            "Continue anyway? (May cause errors)"
        )
        if not response:
            sys.exit(0)

    FLASK_APP = create_flask_app()

    def run_server():
        logger.info(f"Starting Flask server on port {SERVER_PORT}...")
        try:
            FLASK_APP.run(host='localhost', port=SERVER_PORT, debug=False, use_reloader=False, threaded=True)
        except OSError as e:
            logger.error(f"Failed to start server: {e}")

    server_thread = threading.Thread(target=run_server, daemon=True)
    server_thread.start()
    time.sleep(1)  # Wait for server to start
    logger.info("Flask server started")

#
# BROWSER CONTROLLER
#
class BrowserController:
    @staticmethod
    def find_chrome_path():
        """Dynamically find Chrome installation path"""
        if sys.platform == 'win32':
            import winreg

            # Try registry first (most reliable)
            registry_paths = [
                (winreg.HKEY_CURRENT_USER, r'Software\Microsoft\Windows\CurrentVersion\App Paths\chrome.exe'),
                (winreg.HKEY_LOCAL_MACHINE, r'Software\Microsoft\Windows\CurrentVersion\App Paths\chrome.exe'),
                (winreg.HKEY_LOCAL_MACHINE, r'SOFTWARE\Clients\StartMenuInternet\Google Chrome\shell\open\command'),
            ]

            for hkey, reg_path in registry_paths:
                try:
                    key = winreg.OpenKey(hkey, reg_path)
                    try:
                        chrome_path, _ = winreg.QueryValueEx(key, '')
                        # Clean up path (remove quotes and arguments)
                        chrome_path = chrome_path.strip('"').split(' --')[0].split('.exe')[0] + '.exe'
                        if os.path.exists(chrome_path):
                            logger.info(f"Found Chrome via registry: {chrome_path}")
                            return chrome_path
                    finally:
                        winreg.CloseKey(key)
                except (FileNotFoundError, OSError):
                    continue

            # Fallback to common installation paths
            common_paths = [
                r'C:\Program Files\Google\Chrome\Application\chrome.exe',
                r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe',
                os.path.expandvars(r'%LOCALAPPDATA%\Google\Chrome\Application\chrome.exe'),
                os.path.expandvars(r'%PROGRAMFILES%\Google\Chrome\Application\chrome.exe'),
                os.path.expandvars(r'%PROGRAMFILES(X86)%\Google\Chrome\Application\chrome.exe'),
            ]

            for path in common_paths:
                if os.path.exists(path):
                    logger.info(f"Found Chrome at: {path}")
                    return path

        elif sys.platform == 'darwin':
            mac_path = '/Applications/Google Chrome.app/Contents/MacOS/Google Chrome'
            if os.path.exists(mac_path):
                return mac_path
        else:  # Linux
            linux_paths = ['/usr/bin/google-chrome', '/usr/bin/chromium-browser', '/usr/bin/chromium']
            for path in linux_paths:
                if os.path.exists(path):
                    return path

        logger.warning("Chrome not found in any known location")
        return None

    @staticmethod
    def open_vendor_page(vendor_name, part_number):
        """Open vendor page in Chrome"""
        vendor_key = vendor_name.lower().strip()

        if vendor_key not in VENDOR_URLS:
            logger.error(f"Unknown vendor: {vendor_name}")
            return False

        url = VENDOR_URLS[vendor_key].format(part_number)
        logger.info(f"Opening {url} in Chrome...")

        try:
            # Find Chrome dynamically
            chrome_path = BrowserController.find_chrome_path()

            if chrome_path:
                webbrowser.register('chrome', None, webbrowser.BackgroundBrowser(chrome_path))
                webbrowser.get('chrome').open(url)
            else:
                logger.warning("Chrome not found, using default browser")
                webbrowser.open(url)
            return True
        except Exception as e:
            logger.error(f"Failed to open browser: {e}")
            return False
    
    @staticmethod
    def wait_for_scraped_data(timeout=30):
        """Wait for extension to send scraped data"""
        logger.info(f"Waiting up to {timeout}s for scraped data...")
        start_time = time.time()
        
        while time.time() - start_time < timeout:
            try:
                data = DATA_QUEUE.get(timeout=0.5)
                logger.info("Data received from extension")
                return data
            except queue.Empty:
                continue
        
        logger.warning("Timeout waiting for scraped data")
        return None

#
# DATA PROCESSING
#
def sanitize_string(value):
    """Clean string data"""
    if isinstance(value, str):
        return value.replace('\xa0', ' ').replace('\u200e', '').strip()
    return value

def format_date_value(value):
    """Format date value to MM/DD/YYYY string, handling date and datetime objects"""
    if value is None or value == '':
        return None

    # Check if it's a date or datetime object (from Excel or Python)
    if isinstance(value, date):  # This covers both date and datetime
        return value.strftime("%m/%d/%Y")

    # If it's already a string, try to parse and reformat it
    if isinstance(value, str):
        # If it already looks like MM/DD/YYYY, return it
        if len(value) == 10 and value.count('/') == 2:
            try:
                # Validate it's a real date
                datetime.strptime(value, "%m/%d/%Y")
                return value
            except ValueError:
                pass

        # Try to parse datetime string and reformat
        # Handle common formats including "2024-02-26 00:00:00" and "2024-02-26"
        for fmt in ['%Y-%m-%d %H:%M:%S', '%m/%d/%Y %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y']:
            try:
                parsed_date = datetime.strptime(str(value).split('.')[0].strip(), fmt)
                return parsed_date.strftime("%m/%d/%Y")
            except ValueError:
                continue

        # If parsing fails, return original
        return str(value)

    # Fallback to string conversion
    return str(value) if value is not None else None

def format_price_value(value):
    """Format price value to numeric type for Excel, cleaning strings"""
    if value is None or value == '' or value == 'Not Found':
        return None

    # If already numeric, return as-is
    if isinstance(value, (int, float)):
        return float(value)

    # Clean string and convert to float
    try:
        # Remove $, commas, and other non-numeric characters except decimal point
        cleaned = ''.join(c for c in str(value) if c.isdigit() or c == '.')
        return float(cleaned) if cleaned else None
    except (ValueError, AttributeError):
        return None

def prepare_date_for_excel(value):
    """Prepare date value for Excel - returns date object (no time) or None"""
    if value is None or value == '' or value == 'Not Found':
        return None

    # If it's already a date object, return as-is
    if isinstance(value, date) and not isinstance(value, datetime):
        return value

    # If it's a datetime object, strip time component
    if isinstance(value, datetime):
        return value.date()

    # If it's a string, parse it to date
    if isinstance(value, str):
        # Skip invalid values
        if value.strip() in ['', 'Not Found', 'None']:
            return None

        # Try to parse common date formats
        for fmt in ['%m/%d/%Y', '%Y-%m-%d', '%m/%d/%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S']:
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue

    # If we can't parse it, return None
    return None

def clean_value_for_excel(value):
    """Clean value for Excel - convert empty/invalid strings to None to prevent '0' display"""
    if value is None:
        return None
    if value == '' or value == 'None' or value == 'Not Found':
        return None
    if isinstance(value, str) and value.strip() in ['', 'Not Found', 'None']:
        return None
    return value

def prepare_aci_for_excel(value):
    """Ensure ACI # is stored as a number in Excel when it is purely numeric.

    - Returns int for digit-only strings (e.g., "12345" -> 12345)
    - Preserves non-numeric ACI values (e.g., with hyphens/letters)
    - Normalizes floats that are whole numbers to int (e.g., 123.0 -> 123)
    """
    if value is None:
        return None
    try:
        # Already numeric
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value) if value.is_integer() else value

        s = str(value).strip()
        if s.isdigit():
            # Convert digit-only strings to int for proper Excel numeric typing
            return int(s)
    except Exception:
        pass
    return value

def parse_vendor_data(raw_data, vendor_name):
    """Parse raw scraped data into standardized format"""
    if not raw_data:
        return None
    
    vendor_key = vendor_name.lower().strip()
    
    # Extract common fields
    description = raw_data.get('description', 'Not Found')
    price_raw = raw_data.get('price', 'Not Found')
    unit = raw_data.get('unit', 'Not Found')
    mfr_number = raw_data.get('mfrNumber', 'Not Found')
    brand = raw_data.get('brand', 'Not Found')
    
    # Vendor-specific cleanup
    if vendor_key == 'grainger':
        price, unit, qty = cleanup_grainger_data(price_raw, unit, mfr_number)
    elif vendor_key in ['mcmaster', 'mcmaster-carr']:
        price, unit, qty = cleanup_mcmaster_data(price_raw, unit)
        # For McMaster, don't replace MFR number - leave it empty if not found
    elif vendor_key == 'festo':
        price, unit, qty = cleanup_festo_data(price_raw, unit, raw_data.get('qty'))
    elif vendor_key == 'zoro':
        price, unit, qty = cleanup_zoro_data(price_raw, unit)
    else:
        price = price_raw
        qty = 1
    
    return {
        'description': description,
        'price': price,
        'unit': unit,
        'qty': qty,
        'mfr_number': mfr_number,
        'brand': brand
    }

def cleanup_grainger_data(price, unit, mfr):
    """Clean Grainger-specific data"""
    # Remove $ and clean price
    price = price.replace('$', '').strip() if price != "Not Found" else "Not Found"
    
    # Parse unit and quantity
    unit = unit.replace('/', '').strip()
    qty = 1
    
    if " of " in unit:
        parts = unit.split(" of ")
        unit = parts[0].strip()
        try:
            qty = int(parts[1].strip())
        except ValueError:
            qty = 1
    
    # Clean MFR number
    if ' ' in mfr:
        mfr = mfr.split()[-1]
    
    return price, unit, qty

def cleanup_mcmaster_data(price, unit):
    """Clean McMaster-specific data"""
    qty = 1

    # Parse price - handle "$10.33 per pack of 10" format
    if "Not Found" not in price:
        if " per " in price.lower():
            parts = price.lower().split("per", 1)
            price = parts[0].replace('$', '').strip()

            # Extract unit and quantity from the "per" part (e.g., "pack of 10")
            per_part = parts[1].strip()
            if "pack" in per_part and "of" in per_part:
                unit = "pack"
                # Extract quantity number
                qty_parts = per_part.split("of", 1)
                if len(qty_parts) > 1:
                    try:
                        qty = int(''.join(filter(str.isdigit, qty_parts[1])))
                    except ValueError:
                        qty = 1
            elif "each" in per_part:
                unit = "each"
                qty = 1
            # If unit info was in price but couldn't parse, leave unit as-is

        elif "each" in price.lower():
            parts = price.lower().split("each", 1)
            price = parts[0].replace('$', '').strip()
            unit = "each"
            qty = 1
        else:
            price = price.replace('$', '').strip()

    # Only parse unit parameter if it wasn't already extracted from price
    if "Not Found" not in unit and unit not in ["pack", "each"]:
        if "each" in unit.lower():
            unit = "each"
            qty = 1
        elif "pack" in unit.lower() and "of" in unit.lower():
            parts = unit.split(" of ", 1)
            unit = "pack"
            try:
                qty = int(parts[1].strip())
            except ValueError:
                qty = 1

    return price, unit, qty

def cleanup_festo_data(price, unit, qty_raw):
    """Clean Festo-specific data"""
    price = price.replace('$', '').strip() if price != "Not Found" else "Not Found"
    
    qty = 1
    if qty_raw and qty_raw != "Not Found":
        try:
            qty = int(qty_raw)
        except ValueError:
            qty = 1
    
    return price, unit, qty

def cleanup_zoro_data(price, unit):
    """Clean Zoro-specific data"""
    qty = 1
    
    if "Not Found" not in price:
        # Remove newlines and extra text
        price = price.replace('\r\n', ' ').replace('\n', ' ')
        price = price.replace('product price:', '').strip()
        
        # Parse format: "$123.45 / pk 10" or "$45.67 / ea"
        parts = price.split(',')
        
        for part in parts:
            part = part.strip().replace('$', '')
            detail_parts = part.split('/')
            
            if len(detail_parts) >= 2:
                try:
                    price = float(detail_parts[0].strip())
                except ValueError:
                    continue
                
                unit_part = detail_parts[1].strip()
                
                if 'pk' in unit_part:
                    unit = 'pk'
                    try:
                        qty = int(unit_part.split()[1])
                    except:
                        qty = 1
                elif 'ea' in unit_part:
                    unit = 'each'
                    qty = 1
                elif 'pr' in unit_part:
                    unit = 'pair'
                    qty = 2
                
                break
    
    return str(price), unit, qty

def calculate_percentage_change(old_price, new_price):
    """Calculate percentage change between prices"""
    def clean_price(price):
        if price is None or price == '' or price == 'Not Found':
            return None
        if isinstance(price, (int, float)):
            return float(price)
        cleaned = ''.join(c for c in str(price) if c.isdigit() or c == '.')
        return float(cleaned) if cleaned else None
    
    try:
        old = clean_price(old_price)
        new = clean_price(new_price)
        
        if old is None or new is None:
            return None
        
        if old == 0:
            return float('inf') if new != 0 else 0.0
        
        return round(((new - old) / old) * 100, 2)
    except Exception as e:
        logger.error(f"Error calculating percentage change: {e}")
        return None

def update_price_history(entry_data, current_data):
    """Update price history in CSV format"""
    if current_data[14] is None:
        current_data[14] = ""

    if current_data[14] != "":
        entry_data[14] = current_data[14] + f", Date: {current_data[10]} Price: {current_data[9]}"
    else:
        entry_data[14] = f"Date: {current_data[10]} Price: {current_data[9]}"

    entry_data[12] = current_data[9]  # Last updated price
    entry_data[13] = current_data[10]  # Last updated date

    return entry_data

#
# EXCEL OPERATIONS
#
def process_excel(file_path, search_string):
    """Search for ACI number and return row data"""
    workbook = None
    try:
        logger.info(f"Searching Excel for ACI#: {search_string}")
        workbook = load_workbook(file_path, read_only=False, keep_vba=True)
        sheet = workbook["Purchase Parts"]

        for row in sheet.iter_rows(min_row=1, max_col=15, max_row=sheet.max_row):
            cell_value = row[0].value
            # Cast both to string and strip whitespace for comparison
            if cell_value is not None:
                # Handle both numeric and text values
                cell_str = str(sanitize_string(cell_value)).strip()
                search_str = str(search_string).strip()
                if cell_str == search_str:
                    row_index = row[0].row
                    current_data = []

                    # Read and format each cell value
                    for idx, cell in enumerate(row[:15]):
                        value = cell.value

                        # Format dates (fields 10, 13: Date, Last Updated Date)
                        if idx in [10, 13]:
                            value = format_date_value(value)
                        else:
                            value = sanitize_string(value)

                        current_data.append(value)

                    logger.info(f"Found match at row {row_index}")
                    return current_data, row_index

        logger.info("No match found")
        return "NOT_FOUND", None
    except Exception as e:
        logger.error(f"Excel error: {e}")
        return None, None
    finally:
        if workbook:
            workbook.close()

def save_to_excel(file_path, row_index, data):
    """Save updated data to Excel with proper formatting"""
    workbook = None
    try:
        workbook = load_workbook(file_path, read_only=False, keep_vba=True)
        sheet = workbook["Purchase Parts"]

        for idx, value in enumerate(data):
            cell = sheet.cell(row=row_index, column=idx + 1)

            # Format the value based on field type
            formatted_value = value

            # Ensure ACI # (col 0) is numeric when purely digits
            if idx == 0:
                formatted_value = prepare_aci_for_excel(value)
            # Format prices (fields 9, 12: Unit Price, Last Updated Price)
            elif idx in [9, 12]:
                formatted_value = format_price_value(value)
            # Format dates (fields 10, 13: Date, Last Updated Date)
            elif idx in [10, 13]:
                formatted_value = prepare_date_for_excel(value)
            else:
                # For all other fields, clean empty strings to None
                formatted_value = clean_value_for_excel(value)

            # Write the value
            cell.value = formatted_value

            # Apply number format from COLUMN_FORMATS if defined
            if idx in COLUMN_FORMATS:
                cell.number_format = COLUMN_FORMATS[idx]

        workbook.save(filename=file_path)
        logger.info("Excel file saved successfully")
        return True
    except Exception as e:
        logger.error(f"Error saving Excel: {e}")
        return False
    finally:
        if workbook:
            workbook.close()

def add_new_row_to_excel(file_path, aci_number, vendor, vendor_part_number):
    """Add a new row to Excel with basic information, copying formatting from last row"""
    workbook = None
    try:
        workbook = load_workbook(file_path, read_only=False, keep_vba=True)
        sheet = workbook["Purchase Parts"]

        # Find the last row with data and next empty row
        last_row = sheet.max_row
        next_row = last_row + 1

        # Create new entry with ACI#, Vendor, and Vendor Part# (datetime object, not string)
        new_data = [None] * 15
        new_data[0] = prepare_aci_for_excel(aci_number)  # ACI # (numeric if digits only)
        new_data[6] = vendor  # Vendor
        new_data[7] = vendor_part_number  # Vendor Part #
        new_data[10] = datetime.now()  # Date as datetime object

        # Copy formatting from last row and write values
        from copy import copy
        for idx, value in enumerate(new_data):
            col_num = idx + 1

            # Get the cell from the last row to copy formatting from
            source_cell = sheet.cell(row=last_row, column=col_num)
            target_cell = sheet.cell(row=next_row, column=col_num)

            # Copy cell formatting (font, border, fill, alignment, etc.)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)
                # Copy number format from source, will be overridden below if in COLUMN_FORMATS
                target_cell.number_format = copy(source_cell.number_format)

            # Set the value
            target_cell.value = value

            # Override with enforced column formats for critical fields
            if idx in COLUMN_FORMATS:
                target_cell.number_format = COLUMN_FORMATS[idx]

        workbook.save(filename=file_path)
        logger.info(f"New ACI# {aci_number} added at row {next_row} (formatting applied)")

        # Format dates for display (convert datetime objects to strings)
        display_data = new_data.copy()
        for idx in [10, 13]:  # Date fields
            if display_data[idx] is not None:
                display_data[idx] = format_date_value(display_data[idx])

        return display_data, next_row
    except Exception as e:
        logger.error(f"Error adding new row: {e}")
        return None, None
    finally:
        if workbook:
            workbook.close()

def prompt_add_new_aci(aci_number):
    """Ask user if they want to add a new ACI number"""
    response = messagebox.askyesno(
        "ACI Number Not Found",
        f"ACI# {aci_number} not found in the database.\n\n"
        "Would you like to add this as a new entry?"
    )
    return response

def get_new_aci_details(aci_number):
    """Get vendor and vendor part number for new ACI"""
    global _APP_ICON_PHOTO
    root = tk.Tk()
    root.title(f"Add New ACI# - {aci_number}")

    # Set window icon
    try:
        if getattr(sys, 'frozen', False):
            icon_ico = os.path.join(sys._MEIPASS, 'icon.ico')
            icon_png = os.path.join(sys._MEIPASS, 'icon.png')
        else:
            icon_ico = 'icon.ico'
            icon_png = 'icon.png'

        if os.path.exists(icon_ico):
            root.iconbitmap(icon_ico)
        elif os.path.exists(icon_png):
            if _APP_ICON_PHOTO is None:
                _APP_ICON_PHOTO = tk.PhotoImage(file=icon_png)
            root.iconphoto(True, _APP_ICON_PHOTO)
    except Exception as e:
        logger.warning(f"Could not set window icon: {e}")

    # Instruction label
    tk.Label(root, text=f"Adding New ACI#: {aci_number}", font=("Arial", 11, "bold")).pack(pady=(15, 10), padx=20)

    # Vendor dropdown
    tk.Label(root, text="Vendor:", font=("Arial", 10)).pack(pady=(10, 2), padx=20)
    vendor_var = tk.StringVar(root)
    vendor_options = ['Grainger', 'McMaster-Carr', 'Festo', 'Zoro', 'Other']
    vendor_var.set('Other')
    vendor_menu = tk.OptionMenu(root, vendor_var, *vendor_options)
    vendor_menu.config(font=("Arial", 10), width=20)
    vendor_menu.pack(pady=2, padx=20)

    # Vendor Part Number entry
    tk.Label(root, text="Vendor Part Number:", font=("Arial", 10)).pack(pady=(10, 2), padx=20)
    part_entry = tk.Entry(root, font=("Arial", 10), width=25)
    part_entry.pack(pady=2, padx=20)

    result = {'vendor': None, 'part_number': None}

    def on_submit():
        vendor = vendor_var.get()
        part_number = part_entry.get().strip()

        if not part_number:
            messagebox.showwarning("Missing Information", "Please enter a Vendor Part Number")
            return

        result['vendor'] = vendor
        result['part_number'] = part_number
        root.destroy()

    def on_cancel():
        root.destroy()

    part_entry.bind('<Return>', lambda e: on_submit())
    part_entry.bind('<Escape>', lambda e: on_cancel())

    # Buttons
    button_frame = tk.Frame(root)
    button_frame.pack(pady=15, padx=20)

    cancel_btn = tk.Button(button_frame, text="Cancel", command=on_cancel, font=("Arial", 10), bg="#f44336", fg="white", width=12)
    cancel_btn.pack(side=tk.LEFT, padx=5)

    submit_btn = tk.Button(button_frame, text="Add", command=on_submit, font=("Arial", 10), bg="#4CAF50", fg="white", width=12)
    submit_btn.pack(side=tk.LEFT, padx=5)

    root.protocol("WM_DELETE_WINDOW", on_cancel)

    # Center the window after all widgets are added
    root.update_idletasks()
    window_width = root.winfo_width()
    window_height = root.winfo_height()
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    center_x = int((screen_width - window_width) / 2)
    center_y = int((screen_height - window_height) / 2)
    root.geometry(f"+{center_x}+{center_y}")

    root.lift()
    root.attributes('-topmost', True)
    root.after_idle(root.attributes, '-topmost', False)
    root.focus_force()
    part_entry.focus_set()

    root.mainloop()

    return result['vendor'], result['part_number']

#
# GUI COMPONENTS
#
# Global variable to cache PhotoImage to prevent memory leaks
_APP_ICON_PHOTO = None

def get_search_string():
    """Prompt user for ACI number or batch update"""
    global _APP_ICON_PHOTO
    root = tk.Tk()
    root.title("Advantage Conveyor")

    # Set window icon
    try:
        if getattr(sys, 'frozen', False):
            # Running as compiled exe - try ICO first for better Windows support
            icon_ico = os.path.join(sys._MEIPASS, 'icon.ico')
            icon_png = os.path.join(sys._MEIPASS, 'icon.png')
        else:
            # Running as script
            icon_ico = 'icon.ico'
            icon_png = 'icon.png'

        # Try ICO first (native Windows format), fallback to PNG
        if os.path.exists(icon_ico):
            root.iconbitmap(icon_ico)
        elif os.path.exists(icon_png):
            if _APP_ICON_PHOTO is None:
                _APP_ICON_PHOTO = tk.PhotoImage(file=icon_png)
            root.iconphoto(True, _APP_ICON_PHOTO)
    except Exception as e:
        logger.warning(f"Could not set window icon: {e}")

    # Status label
    status_label = tk.Label(root, text="Server Status: Running ✓", font=("Arial", 9), fg="green")
    status_label.pack(pady=(12, 8), padx=20)

    # Instruction label
    instruction_label = tk.Label(root, text="Enter ACI Number:", font=("Arial", 10))
    instruction_label.pack(pady=(8, 5), padx=20)

    # Entry field
    entry = tk.Entry(root, font=("Arial", 10), width=25)
    entry.pack(pady=8, padx=20)

    result = {'value': None, 'mode': 'single'}

    def on_submit():
        result['value'] = entry.get()
        result['mode'] = 'single'
        root.destroy()

    def on_batch():
        result['value'] = 'BATCH_MODE'
        result['mode'] = 'batch'
        root.destroy()

    def on_cancel():
        result['value'] = None
        result['mode'] = 'single'
        root.destroy()

    # Bind Enter key to submit and Escape key to cancel
    entry.bind('<Return>', lambda e: on_submit())
    entry.bind('<Escape>', lambda e: on_cancel())

    # Buttons
    button_frame = tk.Frame(root)
    button_frame.pack(pady=8, padx=20)

    cancel_btn = tk.Button(button_frame, text="Cancel", command=on_cancel, font=("Arial", 10), bg="#f44336", fg="white", width=12)
    cancel_btn.pack(side=tk.LEFT, padx=5)

    submit_btn = tk.Button(button_frame, text="Submit", command=on_submit, font=("Arial", 10), bg="#4CAF50", fg="white", width=12)
    submit_btn.pack(side=tk.LEFT, padx=5)

    # Batch Update button
    batch_btn = tk.Button(root, text="Batch Update", command=on_batch, font=("Arial", 10), bg="#2196F3", fg="white", width=15)
    batch_btn.pack(pady=8, padx=20)

    root.protocol("WM_DELETE_WINDOW", on_cancel)

    # Center the window after all widgets are added
    root.update_idletasks()
    window_width = root.winfo_width()
    window_height = root.winfo_height()
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    center_x = int((screen_width - window_width) / 2)
    center_y = int((screen_height - window_height) / 2)
    root.geometry(f"+{center_x}+{center_y}")

    root.lift()
    root.attributes('-topmost', True)
    root.after_idle(root.attributes, '-topmost', False)
    root.focus_force()
    entry.focus_set()
    entry.icursor(tk.END)

    root.mainloop()

    return result

def compare_and_highlight(widget, current_value, new_value):
    """Highlight widget if values differ"""
    if str(new_value).strip() != str(current_value).strip():
        widget.config(bg="yellow")
    else:
        widget.config(bg="white")

def switch_checkbox_state(index, checkboxes, text_boxes, current_text_boxes, fields, entry_data, current_data):
    """Toggle checkbox and update display"""
    checkbox = checkboxes[index]
    # Get the state AFTER the checkbox has been toggled (it's already changed by the time this is called)
    state = checkbox.var.get()

    if state:  # Checked - use current data
        if fields[index] == "Description":
            text_boxes[index].delete("1.0", tk.END)
            text_boxes[index].insert(tk.END, current_data[index] if current_data[index] is not None else "")
            text_boxes[index].config(bg="white")
        else:
            text_boxes[index].delete(0, tk.END)
            text_boxes[index].insert(0, str(current_data[index]) if current_data[index] is not None else "")
            text_boxes[index].config(bg="white")
    else:  # Unchecked - use entry data
        if fields[index] == "Description":
            text_boxes[index].delete("1.0", tk.END)
            text_boxes[index].insert(tk.END, entry_data[index] if entry_data[index] is not None else "")
            compare_and_highlight(text_boxes[index], current_data[index], entry_data[index])
        else:
            text_boxes[index].delete(0, tk.END)
            text_boxes[index].insert(0, str(entry_data[index]) if entry_data[index] is not None else "")
            compare_and_highlight(text_boxes[index], current_data[index], entry_data[index])

def user_form(current_data, entry_data, fields, file_path, row_index, tab_id=None):
    """Display GUI form for user confirmation"""
    global _APP_ICON_PHOTO
    root = tk.Tk()
    root.title("Update Data - ACI# " + str(current_data[0]))

    # Set window icon
    try:
        if getattr(sys, 'frozen', False):
            # Running as compiled exe - try ICO first for better Windows support
            icon_ico = os.path.join(sys._MEIPASS, 'icon.ico')
            icon_png = os.path.join(sys._MEIPASS, 'icon.png')
        else:
            # Running as script
            icon_ico = 'icon.ico'
            icon_png = 'icon.png'

        # Try ICO first (native Windows format), fallback to PNG
        if os.path.exists(icon_ico):
            root.iconbitmap(icon_ico)
        elif os.path.exists(icon_png):
            if _APP_ICON_PHOTO is None:
                _APP_ICON_PHOTO = tk.PhotoImage(file=icon_png)
            root.iconphoto(True, _APP_ICON_PHOTO)
    except Exception as e:
        logger.warning(f"Could not set window icon: {e}")

    def on_closing():
        root.quit()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_closing)

    large_font = tkFont.Font(family="Arial", size=10)

    # Headers
    tk.Label(root, text="Entry Data", font=large_font).grid(row=0, column=1, padx=5, pady=5)
    tk.Label(root, text="Current Data", font=large_font).grid(row=0, column=2, padx=2, pady=5)
    tk.Label(root, text="KEEP", font=large_font).grid(row=0, column=3, padx=(2, 10), pady=2, sticky="w")

    text_boxes = []
    current_text_boxes = []
    checkboxes = []

    for i, field in enumerate(fields):
        tk.Label(root, text=field, font=large_font).grid(row=i + 1, column=0, padx=5, pady=5, sticky="e")

        is_not_found = (entry_data[i] == "Not Found")

        if field == "Description":
            # Multi-line text for description
            text_box = tk.Text(root, font=large_font, height=7, width=40, wrap="word")
            text_box.grid(row=i + 1, column=1, padx=5, pady=5)
            entry_text = "" if is_not_found or entry_data[i] is None else str(entry_data[i])
            text_box.insert(tk.END, entry_text)

            current_text_box = tk.Text(root, font=large_font, height=7, width=40, wrap="word", state="normal", takefocus=0)
            current_text_box.grid(row=i + 1, column=2, padx=50, pady=5)
            current_text = "" if current_data[i] is None else str(current_data[i])
            current_text_box.insert(tk.END, current_text)

            if entry_text.strip() != current_text.strip():
                text_box.config(bg="yellow")
        else:
            # Single-line entry
            text_box = tk.Entry(root, font=large_font, width=40)
            text_box.grid(row=i + 1, column=1, padx=5, pady=5)
            entry_text = "" if is_not_found or entry_data[i] is None else str(entry_data[i])
            text_box.insert(0, entry_text)

            current_text_box = tk.Entry(root, font=large_font, state='normal', width=40, takefocus=0)
            current_text_box.grid(row=i + 1, column=2, padx=50, pady=5)
            current_text = "" if current_data[i] is None else str(current_data[i])
            current_text_box.insert(0, current_text)
            current_text_box.config(state="readonly")

        if is_not_found:
            text_box.config(bg="red")

        compare_and_highlight(text_box, entry_data[i], current_data[i])

        text_boxes.append(text_box)
        current_text_boxes.append(current_text_box)

        # Checkbox
        var = tk.BooleanVar()
        checkbox = tk.Checkbutton(root, text="", variable=var, onvalue=True, offvalue=False, takefocus=0)
        checkbox.var = var
        checkbox.config(command=lambda idx=i: switch_checkbox_state(
            idx, checkboxes, text_boxes, current_text_boxes, fields, entry_data, current_data
        ))
        checkbox.grid(row=i + 1, column=3, padx=(2, 10), pady=2, sticky="w")
        checkboxes.append(checkbox)

    def submit():
        try:
            # Update entry_data based on checkboxes
            for i, field in enumerate(fields):
                checkbox_state = checkboxes[i].var.get()
                if checkbox_state:
                    # Use current data
                    if field == "Description":
                        entry_data[i] = current_text_boxes[i].get("1.0", tk.END).strip()
                    else:
                        entry_data[i] = current_text_boxes[i].get()
                else:
                    # Use entry data
                    if field == "Description":
                        entry_data[i] = text_boxes[i].get("1.0", tk.END).strip()
                    else:
                        entry_data[i] = text_boxes[i].get()

            # Calculate price change and update date
            percent_change = 0
            if current_data[9] not in ['Legacy', 'None', None] and entry_data[9] not in ['Legacy', 'None', None]:
                try:
                    percent_change = calculate_percentage_change(current_data[9], entry_data[9])

                    # Update entry_data with calculated values
                    entry_data[10] = datetime.now()  # Date
                    entry_data[11] = percent_change  # Change %

                    if percent_change and abs(percent_change) >= 1:
                        update_price_history(entry_data, current_data)
                except ValueError as e:
                    logger.error(f"Price conversion error: {e}")
            else:
                # Even if we can't calculate percentage change, update the date
                entry_data[10] = datetime.now()

            # Alert on significant price changes
            if percent_change and abs(percent_change) >= 20:
                approve = messagebox.askyesno(
                    "Price Change Alert",
                    f"Price change is significant: {percent_change}%. Proceed?"
                )
                if not approve:
                    return
            elif percent_change and percent_change <= -10:
                approve = messagebox.askyesno(
                    "Price Change Alert",
                    f"Price decrease: {percent_change}%. Proceed?"
                )
                if not approve:
                    return

            # Save to Excel
            if save_to_excel(file_path, row_index, entry_data):
                # Signal extension to close the tab
                if tab_id:
                    TABS_TO_CLOSE.add(tab_id)
                    logger.info(f"Added tab {tab_id} to close queue")
                root.destroy()
            else:
                messagebox.showerror("Error", "Failed to save data")

        except Exception as e:
            logger.error(f"Error in submit: {e}")
            messagebox.showerror("Error", str(e))

    def cancel():
        # Signal extension to close the tab
        if tab_id:
            TABS_TO_CLOSE.add(tab_id)
            logger.info(f"Added tab {tab_id} to close queue")
        root.quit()
        root.destroy()

    # Check All button and action buttons on same row
    check_all_state = {'checked': False}

    def toggle_all_checkboxes():
        check_all_state['checked'] = not check_all_state['checked']
        for checkbox in checkboxes:
            checkbox.var.set(check_all_state['checked'])
            # Trigger the visual update by updating the display
            idx = checkboxes.index(checkbox)
            if check_all_state['checked']:
                # Use current data
                if fields[idx] == "Description":
                    text_boxes[idx].delete("1.0", tk.END)
                    text_boxes[idx].insert(tk.END, current_data[idx] if current_data[idx] is not None else "")
                    text_boxes[idx].config(bg="white")
                else:
                    text_boxes[idx].delete(0, tk.END)
                    text_boxes[idx].insert(0, str(current_data[idx]) if current_data[idx] is not None else "")
                    text_boxes[idx].config(bg="white")
            else:
                # Use entry data
                if fields[idx] == "Description":
                    text_boxes[idx].delete("1.0", tk.END)
                    text_boxes[idx].insert(tk.END, entry_data[idx] if entry_data[idx] is not None else "")
                    compare_and_highlight(text_boxes[idx], current_data[idx], entry_data[idx])
                else:
                    text_boxes[idx].delete(0, tk.END)
                    text_boxes[idx].insert(0, str(entry_data[idx]) if entry_data[idx] is not None else "")
                    compare_and_highlight(text_boxes[idx], current_data[idx], entry_data[idx])

        # Update button text
        if check_all_state['checked']:
            check_all_button.config(text="Uncheck All")
        else:
            check_all_button.config(text="Check All")

    check_all_button = tk.Button(
        root,
        text="Check All",
        command=toggle_all_checkboxes,
        font=large_font,
        takefocus=0,
        width=10
    )
    check_all_button.grid(row=len(fields) + 1, column=2, padx=10, pady=10, sticky="e")

    cancel_button = tk.Button(root, text="Cancel", command=cancel, font=large_font, bg="#f44336", fg="white", takefocus=0)
    cancel_button.grid(row=len(fields) + 1, column=1, padx=20, pady=10)

    submit_button = tk.Button(root, text="Submit", command=submit, font=large_font, bg="#4CAF50", fg="white", takefocus=0)
    submit_button.grid(row=len(fields) + 1, column=2, padx=10, pady=10)

    # Bind keyboard shortcuts
    root.bind('<Control-s>', lambda e: submit())
    root.bind('<Control-S>', lambda e: submit())
    root.bind('<Control-x>', lambda e: cancel())
    root.bind('<Control-X>', lambda e: cancel())

    root.protocol("WM_DELETE_WINDOW", cancel)
    root.transient()
    root.grab_set()

    # Center the window after all widgets are added
    root.update_idletasks()
    window_width = root.winfo_width()
    window_height = root.winfo_height()
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    center_x = int((screen_width - window_width) / 2)
    center_y = int((screen_height - window_height) / 2)
    root.geometry(f"+{center_x}+{center_y}")

    # Make window active and bring to front
    root.lift()
    root.attributes('-topmost', True)
    root.after_idle(root.attributes, '-topmost', False)
    root.focus_force()

    root.mainloop()

#
# MAIN WORKFLOW
#
def process_item(vendor_name, part_number, current_data, entry_data):
    """Orchestrate the scraping workflow"""
    logger.info(f"Processing {vendor_name} part {part_number}")

    # Clear any stale data before opening new browser tab
    clear_stale_data()

    # Open browser
    if not BrowserController.open_vendor_page(vendor_name, part_number):
        logger.error("Failed to open browser")
        return None

    # Wait for scraped data (reduced timeout for faster workflow)
    raw_data = BrowserController.wait_for_scraped_data(timeout=15)

    if not raw_data:
        logger.warning("No data received from extension")
        messagebox.showwarning("Timeout", "No data received. Extension may not be installed or page took too long.")
        return None

    # Extract tab ID
    tab_id = raw_data.get('tabId')

    # Parse data
    parsed_data = parse_vendor_data(raw_data, vendor_name)

    if not parsed_data:
        logger.error("Failed to parse vendor data")
        return None

    # Update entry_data
    entry_data[1] = parsed_data['mfr_number']
    entry_data[2] = parsed_data['brand']
    entry_data[3] = parsed_data['description']
    entry_data[4] = parsed_data['qty']
    entry_data[5] = parsed_data['unit']
    entry_data[9] = parsed_data['price']
    entry_data[10] = datetime.now().strftime("%m/%d/%Y")  # Date as string for display
    entry_data[11] = calculate_percentage_change(current_data[9], parsed_data['price'])  # Change %

    logger.info(f"Data parsed successfully: Price=${parsed_data['price']}, Qty={parsed_data['qty']}")
    return tab_id

def is_vendor_auto(vendor):
    """Check if vendor supports automatic scraping"""
    return vendor.lower() in ['grainger', 'mcmaster-carr', 'mcmaster', 'zoro', 'festo']

#
# BATCH UPDATE FUNCTIONS
#
def batch_update_dialog():
    """Dialog to get list of ACI numbers for batch update"""
    global _APP_ICON_PHOTO
    root = tk.Tk()
    root.title("Batch Update - Enter ACI Numbers")

    # Set window icon
    try:
        if getattr(sys, 'frozen', False):
            icon_ico = os.path.join(sys._MEIPASS, 'icon.ico')
            icon_png = os.path.join(sys._MEIPASS, 'icon.png')
        else:
            icon_ico = 'icon.ico'
            icon_png = 'icon.png'

        if os.path.exists(icon_ico):
            root.iconbitmap(icon_ico)
        elif os.path.exists(icon_png):
            if _APP_ICON_PHOTO is None:
                _APP_ICON_PHOTO = tk.PhotoImage(file=icon_png)
            root.iconphoto(True, _APP_ICON_PHOTO)
    except Exception as e:
        logger.warning(f"Could not set window icon: {e}")

    # Instructions
    tk.Label(root, text="Batch Update", font=("Arial", 12, "bold")).pack(pady=(15, 5))
    tk.Label(root, text="Enter ACI numbers (one per line or comma-separated):", font=("Arial", 10)).pack(pady=5)

    # Text area
    text_area = scrolledtext.ScrolledText(root, font=("Arial", 10), width=50, height=15)
    text_area.pack(pady=10, padx=20)

    result = {'aci_list': None}

    def on_submit():
        text = text_area.get("1.0", tk.END).strip()
        if not text:
            messagebox.showwarning("Empty List", "Please enter at least one ACI number")
            return

        # Parse ACI numbers (support both newlines and commas)
        aci_list = []
        for line in text.split('\n'):
            for item in line.split(','):
                aci = item.strip()
                if aci:
                    aci_list.append(aci.upper())

        if not aci_list:
            messagebox.showwarning("Empty List", "No valid ACI numbers found")
            return

        result['aci_list'] = aci_list
        root.destroy()

    def on_cancel():
        root.destroy()

    # Buttons
    button_frame = tk.Frame(root)
    button_frame.pack(pady=10)

    cancel_btn = tk.Button(button_frame, text="Cancel", command=on_cancel, font=("Arial", 10), bg="#f44336", fg="white", width=12)
    cancel_btn.pack(side=tk.LEFT, padx=5)

    submit_btn = tk.Button(button_frame, text="Start Batch", command=on_submit, font=("Arial", 10), bg="#4CAF50", fg="white", width=12)
    submit_btn.pack(side=tk.LEFT, padx=5)

    root.protocol("WM_DELETE_WINDOW", on_cancel)

    # Center the window after all widgets are added
    root.update_idletasks()
    window_width = root.winfo_width()
    window_height = root.winfo_height()
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    center_x = int((screen_width - window_width) / 2)
    center_y = int((screen_height - window_height) / 2)
    root.geometry(f"+{center_x}+{center_y}")

    root.lift()
    root.attributes('-topmost', True)
    root.after_idle(root.attributes, '-topmost', False)
    root.focus_force()
    text_area.focus_set()

    root.mainloop()

    return result['aci_list']

def validate_batch_match(current_data, scraped_data, vendor_name):
    """Validate if scraped data matches current data for batch update"""
    vendor_key = vendor_name.lower().strip()

    # Check part number match (skip for McMaster as their part numbers may differ from MFR)
    if vendor_key not in ['mcmaster', 'mcmaster-carr']:
        current_part = str(current_data[1]).strip() if current_data[1] else ""
        scraped_part = str(scraped_data.get('mfr_number', '')).strip()

        # Allow some flexibility in part number matching
        part_match = current_part.lower() == scraped_part.lower() if current_part and scraped_part != "Not Found" else True

        if current_part and scraped_part != "Not Found" and not part_match:
            return False, "Part number mismatch"

    # Check unit match
    current_unit = str(current_data[5]).strip().lower() if current_data[5] else ""
    scraped_unit = str(scraped_data.get('unit', '')).strip().lower()

    unit_match = current_unit == scraped_unit if current_unit and scraped_unit != "not found" else True

    if current_unit and scraped_unit != "not found" and not unit_match:
        return False, "Unit mismatch"

    return True, "Match"

def batch_update_worker(file_path, aci_list):
    """Process batch update for list of ACI numbers"""
    results = {
        'updated': [],
        'skipped': [],
        'errors': [],
        'not_found': []
    }

    total = len(aci_list)
    logger.info(f"Starting batch update for {total} ACI numbers")

    for idx, aci in enumerate(aci_list):
        logger.info(f"Processing {idx + 1}/{total}: {aci}")

        try:
            # Look up ACI in Excel
            current_data, row_index = process_excel(file_path, aci)

            if current_data == "NOT_FOUND":
                results['not_found'].append(aci)
                logger.info(f"  ACI {aci} not found in Excel")
                continue

            if current_data is None:
                results['errors'].append((aci, "Excel error"))
                logger.error(f"  Error loading ACI {aci}")
                continue

            # Check if vendor is auto-supported
            vendor_name = current_data[6]
            if not is_vendor_auto(vendor_name):
                results['skipped'].append((aci, f"Manual vendor: {vendor_name}"))
                logger.info(f"  Skipped {aci} - manual vendor")
                continue

            # For hyphenated ACI numbers, only skip for McMaster vendors
            try:
                vendor_key = str(vendor_name).strip().lower() if vendor_name else ""
                if '-' in str(aci) and vendor_key in ['mcmaster', 'mcmaster-carr']:
                    results['skipped'].append((aci, "Hyphenated ACI - McMaster requires manual update"))
                    logger.info(f"  Skipped {aci} - hyphenated ACI for McMaster")
                    continue
            except Exception as e:
                logger.warning(f"  Warning while evaluating hyphen rule for {aci}: {e}")

            # Scrape data
            part_number = current_data[7]
            entry_data = current_data.copy()

            # Clear stale data before each scrape
            clear_stale_data()

            logger.info(f"  Opening {vendor_name} page for {part_number}")
            if not BrowserController.open_vendor_page(vendor_name, part_number):
                results['errors'].append((aci, "Failed to open browser"))
                continue

            raw_data = BrowserController.wait_for_scraped_data(timeout=15)

            if not raw_data:
                results['errors'].append((aci, "Scraping timeout"))
                logger.warning(f"  Timeout for {aci}")
                # Close any registered tabs
                if REGISTERED_TABS:
                    most_recent = max(REGISTERED_TABS.items(), key=lambda x: x[1]['timestamp'])
                    tab_id = most_recent[0]
                    TABS_TO_CLOSE.add(tab_id)
                continue

            tab_id = raw_data.get('tabId')
            parsed_data = parse_vendor_data(raw_data, vendor_name)

            if not parsed_data:
                results['errors'].append((aci, "Failed to parse data"))
                if tab_id:
                    TABS_TO_CLOSE.add(tab_id)
                continue

            # Validate match
            is_match, match_reason = validate_batch_match(current_data, parsed_data, vendor_name)

            if not is_match:
                results['skipped'].append((aci, match_reason))
                logger.info(f"  Skipped {aci} - {match_reason}")
                if tab_id:
                    TABS_TO_CLOSE.add(tab_id)
                continue

            # Check price change within ±15%
            old_price = current_data[9]
            new_price = parsed_data['price']

            if new_price == "Not Found":
                results['skipped'].append((aci, "Price not found"))
                if tab_id:
                    TABS_TO_CLOSE.add(tab_id)
                continue

            percent_change = calculate_percentage_change(old_price, new_price)

            if percent_change is None or abs(percent_change) > 15:
                reason = f"Price change {percent_change}% exceeds ±15%"
                results['skipped'].append((aci, reason))
                logger.info(f"  Skipped {aci} - {reason}")
                if tab_id:
                    TABS_TO_CLOSE.add(tab_id)
                continue

            # Update entry_data
            entry_data[9] = new_price
            entry_data[10] = datetime.now()  # Date
            entry_data[11] = percent_change  # Change %

            if percent_change and abs(percent_change) >= 1:
                update_price_history(entry_data, current_data)

            # Save to Excel
            if save_to_excel(file_path, row_index, entry_data):
                results['updated'].append((aci, f"{old_price} → {new_price} ({percent_change:+.1f}%)"))
                logger.info(f"  Updated {aci}: {old_price} → {new_price} ({percent_change:+.1f}%)")
            else:
                results['errors'].append((aci, "Failed to save"))

            # Close tab
            if tab_id:
                TABS_TO_CLOSE.add(tab_id)

            # Small delay between items
            time.sleep(0.5)

        except Exception as e:
            results['errors'].append((aci, str(e)))
            logger.error(f"  Error processing {aci}: {e}", exc_info=True)

    return results

def show_batch_summary(results):
    """Display batch update summary"""
    global _APP_ICON_PHOTO
    root = tk.Tk()
    root.title("Batch Update Summary")

    # Set window icon
    try:
        if getattr(sys, 'frozen', False):
            icon_ico = os.path.join(sys._MEIPASS, 'icon.ico')
            icon_png = os.path.join(sys._MEIPASS, 'icon.png')
        else:
            icon_ico = 'icon.ico'
            icon_png = 'icon.png'

        if os.path.exists(icon_ico):
            root.iconbitmap(icon_ico)
        elif os.path.exists(icon_png):
            if _APP_ICON_PHOTO is None:
                _APP_ICON_PHOTO = tk.PhotoImage(file=icon_png)
            root.iconphoto(True, _APP_ICON_PHOTO)
    except Exception as e:
        logger.warning(f"Could not set window icon: {e}")

    # Title
    tk.Label(root, text="Batch Update Complete", font=("Arial", 14, "bold")).pack(pady=(15, 10))

    # Summary counts
    summary_frame = tk.Frame(root)
    summary_frame.pack(pady=10)

    tk.Label(summary_frame, text=f"✓ Updated: {len(results['updated'])}", font=("Arial", 11), fg="green").grid(row=0, column=0, padx=15)
    tk.Label(summary_frame, text=f"⊘ Skipped: {len(results['skipped'])}", font=("Arial", 11), fg="orange").grid(row=0, column=1, padx=15)
    tk.Label(summary_frame, text=f"✗ Errors: {len(results['errors'])}", font=("Arial", 11), fg="red").grid(row=0, column=2, padx=15)
    tk.Label(summary_frame, text=f"? Not Found: {len(results['not_found'])}", font=("Arial", 11), fg="gray").grid(row=0, column=3, padx=15)

    # Details text area
    text_area = scrolledtext.ScrolledText(root, font=("Courier", 9), width=85, height=20)
    text_area.pack(pady=10, padx=20)

    # Build details
    details = []

    if results['updated']:
        details.append("=" * 80)
        details.append("UPDATED:")
        details.append("=" * 80)
        for aci, info in results['updated']:
            details.append(f"  {aci}: {info}")

    if results['skipped']:
        details.append("\n" + "=" * 80)
        details.append("SKIPPED:")
        details.append("=" * 80)
        for aci, reason in results['skipped']:
            details.append(f"  {aci}: {reason}")

    if results['errors']:
        details.append("\n" + "=" * 80)
        details.append("ERRORS:")
        details.append("=" * 80)
        for aci, error in results['errors']:
            details.append(f"  {aci}: {error}")

    if results['not_found']:
        details.append("\n" + "=" * 80)
        details.append("NOT FOUND:")
        details.append("=" * 80)
        for aci in results['not_found']:
            details.append(f"  {aci}")

    text_area.insert("1.0", "\n".join(details))
    text_area.config(state="disabled")

    # Close button
    close_btn = tk.Button(root, text="Close", command=root.destroy, font=("Arial", 11), bg="#2196F3", fg="white", width=15)
    close_btn.pack(pady=15)

    root.protocol("WM_DELETE_WINDOW", root.destroy)

    # Center the window after all widgets are added
    root.update_idletasks()
    window_width = root.winfo_width()
    window_height = root.winfo_height()
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    center_x = int((screen_width - window_width) / 2)
    center_y = int((screen_height - window_height) / 2)
    root.geometry(f"+{center_x}+{center_y}")

    root.lift()
    root.attributes('-topmost', True)
    root.after_idle(root.attributes, '-topmost', False)
    root.focus_force()

    root.mainloop()

def main_loop(file_path):
    """Main application loop"""
    logger.info("Starting main application loop")

    while True:
        result = get_search_string()

        if result['value'] is None:
            logger.info("User cancelled, exiting")
            break

        # Check if batch mode
        if result['mode'] == 'batch':
            logger.info("Batch update mode selected")

            # Get list of ACI numbers
            aci_list = batch_update_dialog()

            if not aci_list:
                logger.info("Batch update cancelled")
                continue

            # Process batch
            logger.info(f"Starting batch update for {len(aci_list)} ACIs")
            batch_results = batch_update_worker(file_path, aci_list)

            # Show summary
            show_batch_summary(batch_results)
            logger.info("Batch update completed")
            continue

        # Single mode
        search_string = sanitize_string(result['value']).upper()
        logger.info(f"Searching for: {search_string}")

        try:
            # Search Excel
            result_data = process_excel(file_path, search_string)

            if result_data is None or result_data[0] is None:
                continue

            # Handle "NOT_FOUND" case - offer to add new ACI
            if result_data[0] == "NOT_FOUND":
                logger.info(f"ACI {search_string} not found")

                if prompt_add_new_aci(search_string):
                    # Get vendor and part number
                    vendor, part_number = get_new_aci_details(search_string)

                    if not vendor or not part_number:
                        logger.info("User cancelled adding new ACI")
                        continue

                    # Add new row to Excel
                    new_data, new_row_index = add_new_row_to_excel(file_path, search_string, vendor, part_number)

                    if not new_data:
                        messagebox.showerror("Error", "Failed to add new ACI to Excel")
                        continue

                    # Use the new data as current data
                    current_data = new_data
                    row_index = new_row_index
                    entry_data = current_data.copy()

                    # If auto vendor, try to scrape
                    tab_id = None
                    if is_vendor_auto(vendor):
                        logger.info(f"Auto-scraping enabled for new entry: {vendor}")
                        tab_id = process_item(vendor, part_number, current_data, entry_data)

                        # Check for registered tabs if scraping timed out
                        if tab_id is None and REGISTERED_TABS:
                            most_recent = max(REGISTERED_TABS.items(), key=lambda x: x[1]['timestamp'])
                            tab_id = most_recent[0]
                            logger.info(f"Using registered tab {tab_id}")
                    else:
                        logger.info(f"Manual vendor for new entry: {vendor}")

                    # Show user form
                    user_form(current_data, entry_data, FIELDS, file_path, row_index, tab_id)
                else:
                    logger.info("User declined to add new ACI")
                continue

            # Regular flow - ACI found in Excel
            current_data, row_index = result_data
            entry_data = current_data.copy()

            # Check if vendor supports auto-scraping
            tab_id = None
            vendor_name = current_data[6]
            if is_vendor_auto(vendor_name):
                part_number = current_data[7]
                logger.info(f"Auto-scraping enabled for {vendor_name}")
                tab_id = process_item(vendor_name, part_number, current_data, entry_data)

                # Even if scraping timed out, check if a tab was registered
                if tab_id is None and REGISTERED_TABS:
                    # Get the most recently registered tab (likely the one we just opened)
                    most_recent = max(REGISTERED_TABS.items(), key=lambda x: x[1]['timestamp'])
                    tab_id = most_recent[0]
                    logger.info(f"Using registered tab {tab_id} (scraping may have timed out)")
            else:
                logger.info(f"Manual vendor: {vendor_name}")

            # Show user form (pass tab_id if available)
            user_form(current_data, entry_data, FIELDS, file_path, row_index, tab_id)

        except Exception as e:
            logger.error(f"Error in main loop: {e}", exc_info=True)
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            continue

#
# APPLICATION ENTRY POINT
#
def main():
    """Application entry point"""
    # Parse command-line arguments
    parser = argparse.ArgumentParser(
        description='Multi-Vendor Price Scraper - Auto-update pricing data from multiple suppliers',
        epilog='If no arguments provided, runs in interactive GUI mode'
    )
    parser.add_argument(
        '--batch',
        type=str,
        help='Batch update mode: comma-separated list of ACI numbers (e.g., "ACI001,ACI002,ACI003")'
    )
    parser.add_argument(
        '--batch-file',
        type=str,
        help='Batch update mode: path to file containing ACI numbers (one per line)'
    )

    args = parser.parse_args()

    try:
        # Determine file path - check multiple extensions
        server_file_paths = [
            r'Z:\ACOD\MMLV2.xlsm',
            r'Z:\ACOD\MMLV2.xlsx',
            r'Z:\ACOD\MML.xlsm',
            r'Z:\ACOD\MML.xlsx'
        ]
        local_file_paths = ['MML.xlsm', 'MML.xlsx', 'MMLV2.xlsm', 'MMLV2.xlsx']

        file_path = None

        # Check server paths first
        for path in server_file_paths:
            if os.path.exists(path):
                file_path = path
                logger.info(f"Using server file: {file_path}")
                break

        # If no server file found, check local paths
        if not file_path:
            for path in local_file_paths:
                if os.path.exists(path):
                    file_path = path
                    logger.info(f"Using local file: {file_path}")
                    # Only show warning in GUI mode
                    if not args.batch and not args.batch_file:
                        messagebox.showwarning("Warning", "Server file not found. Using local file.")
                    else:
                        logger.warning("Server file not found. Using local file.")
                    break

        # If still no file found, show error and exit
        if not file_path:
            error_msg = (
                "Excel file not found!\n\n"
                "Please ensure one of the following exists:\n"
                "- Z:\\ACOD\\MMLV2.xlsm (or .xlsx)\n"
                "- MML.xlsm (or .xlsx) in the program directory"
            )
            if not args.batch and not args.batch_file:
                messagebox.showerror("Error", error_msg)
            logger.error("No Excel file found")
            sys.exit(1)

        # Start Flask server
        start_flask_server()

        # Check if batch mode requested via command line
        if args.batch or args.batch_file:
            logger.info("Batch mode activated via command line")

            # Parse ACI list
            aci_list = []
            if args.batch:
                # Parse comma-separated list
                for item in args.batch.split(','):
                    aci = item.strip().upper()
                    if aci:
                        aci_list.append(aci)
                logger.info(f"Batch list from --batch: {len(aci_list)} ACIs")

            elif args.batch_file:
                # Read from file
                if not os.path.exists(args.batch_file):
                    logger.error(f"Batch file not found: {args.batch_file}")
                    print(f"ERROR: File not found: {args.batch_file}")
                    sys.exit(1)

                try:
                    with open(args.batch_file, 'r') as f:
                        for line in f:
                            aci = line.strip().upper()
                            if aci and not aci.startswith('#'):  # Allow comments
                                aci_list.append(aci)
                    logger.info(f"Batch list from file: {len(aci_list)} ACIs")
                except Exception as e:
                    logger.error(f"Failed to read batch file: {e}")
                    print(f"ERROR: Failed to read file: {e}")
                    sys.exit(1)

            if not aci_list:
                logger.error("No ACI numbers found in batch input")
                print("ERROR: No ACI numbers found in batch input")
                sys.exit(1)

            # Run batch update
            logger.info(f"Starting CLI batch update for {len(aci_list)} ACIs")
            print(f"Starting batch update for {len(aci_list)} ACI numbers...")

            batch_results = batch_update_worker(file_path, aci_list)

            # Print results to console
            print("\n" + "=" * 80)
            print("BATCH UPDATE COMPLETE")
            print("=" * 80)
            print(f"✓ Updated:   {len(batch_results['updated'])}")
            print(f"⊘ Skipped:   {len(batch_results['skipped'])}")
            print(f"✗ Errors:    {len(batch_results['errors'])}")
            print(f"? Not Found: {len(batch_results['not_found'])}")
            print("=" * 80)

            if batch_results['updated']:
                print("\nUPDATED:")
                for aci, info in batch_results['updated']:
                    print(f"  {aci}: {info}")

            if batch_results['skipped']:
                print("\nSKIPPED:")
                for aci, reason in batch_results['skipped']:
                    print(f"  {aci}: {reason}")

            if batch_results['errors']:
                print("\nERRORS:")
                for aci, error in batch_results['errors']:
                    print(f"  {aci}: {error}")

            if batch_results['not_found']:
                print("\nNOT FOUND:")
                for aci in batch_results['not_found']:
                    print(f"  {aci}")

            print("\n" + "=" * 80)
            logger.info("CLI batch update completed")

            # Exit after batch processing
            sys.exit(0)

        # Normal GUI mode - Start main loop
        main_loop(file_path)

    except KeyboardInterrupt:
        logger.info("Program terminated by user")
        sys.exit(0)
    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        messagebox.showerror("Fatal Error", str(e))
        sys.exit(1)
    finally:
        logger.info("Cleaning up and exiting...")
        # Daemon threads will automatically terminate when main thread exits

if __name__ == "__main__":
    main()
